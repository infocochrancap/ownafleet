#!/usr/bin/env python3
"""
One-shot importer for Josh's pre-website lead data.

Reads two xlsx files:
  1. COI sheet — 8 funded buyers from 2025
  2. Lead Tracker — active 2026 pipeline + various historicals

Maps each row to the simplified `leads` schema (post migration 012) and
upserts via Supabase REST API. Idempotent by name+company match — re-running
won't create duplicates.

Usage:
    SUPABASE_URL=https://lkfaemhhdxjaqggvlotv.supabase.co \\
    SUPABASE_SECRET_KEY=<service_role_key> \\
    python3 web/scripts/import-spreadsheet-leads.py

The SUPABASE_SECRET_KEY is your service-role key from Supabase project
settings -> API. Required because we're writing as an admin-level caller
(RLS would block a regular user from inserting on behalf of others).
"""

import os
import sys
import json
import urllib.request
import urllib.parse
from datetime import datetime
import openpyxl

# -----------------------------------------------------------------------------
# Config
# -----------------------------------------------------------------------------

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_SECRET = os.environ.get('SUPABASE_SECRET_KEY')
if not SUPABASE_URL or not SUPABASE_SECRET:
    print('ERROR: SUPABASE_URL and SUPABASE_SECRET_KEY must be set in env.')
    sys.exit(1)

COI_FILE = '/Users/joshcochran/Downloads/Equipment - COI - Center of Influence.xlsx'
LEAD_FILE = '/Users/joshcochran/Downloads/Lead Tracker - Cochran Capital - Armada (1).xlsx'

# Map spreadsheet status text -> lead_status enum value
# Anything not in this table maps to 'new' (and prints a warning).
STATUS_MAP = {
    # Closing / late-stage
    'closing': 'terms_accepted',
    'out with lenders': 'full_app_submitted',
    'bank approved': 'approved',

    # In-progress
    'awaiting update from allondra': 'mini_app_submitted',
    'application done, and documents have been uploaded.': 'mini_app_submitted',
    'no documents uploaded yet': 'application_sent',
    '1st meet complete': 'contacted',
    'intro email sent': 'contacted',
    'awaiting equipment assignment': 'terms_accepted',
    'awaiting lender': 'full_app_submitted',

    # Not-now bucket
    'not interested this year': 'not_now',
    'ready for 2026': 'not_now',
    'wants 2026': 'not_now',

    # Archived bucket
    'not at fit': 'archived',
    "won't do it": 'archived',
    'andrew declinced': 'archived',
    'not a fit': 'archived',
}


def http_post(path, payload, prefer='return=representation'):
    """POST to Supabase REST API. Returns parsed JSON response (list)."""
    url = f'{SUPABASE_URL}/rest/v1/{path}'
    body = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=body, method='POST', headers={
        'apikey': SUPABASE_SECRET,
        'Authorization': f'Bearer {SUPABASE_SECRET}',
        'Content-Type': 'application/json',
        'Prefer': prefer,
    })
    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode('utf-8')
            return json.loads(text) if text else []
    except urllib.error.HTTPError as e:
        err = e.read().decode('utf-8') if e.fp else ''
        print(f'  HTTP {e.code}: {err}')
        return None


def http_get(path):
    """GET from Supabase REST API. Returns parsed JSON response (list)."""
    url = f'{SUPABASE_URL}/rest/v1/{path}'
    req = urllib.request.Request(url, method='GET', headers={
        'apikey': SUPABASE_SECRET,
        'Authorization': f'Bearer {SUPABASE_SECRET}',
    })
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode('utf-8'))


def split_name(full):
    """Returns (first, last) from a 'First Last' or 'First Middle Last' string."""
    if not full:
        return None, None
    parts = full.strip().split()
    if len(parts) == 0:
        return None, None
    if len(parts) == 1:
        return parts[0], None
    return parts[0], ' '.join(parts[1:])


def lead_exists(first_name, last_name, company):
    """Check if a lead with this name (or name+company) already exists."""
    if not first_name and not last_name and not company:
        return False
    filters = []
    if first_name:
        filters.append(f'first_name=ilike.{urllib.parse.quote(first_name)}')
    if last_name:
        filters.append(f'last_name=ilike.{urllib.parse.quote(last_name)}')
    if not filters:
        # Fall back to company match
        filters.append(f'company=ilike.{urllib.parse.quote(company)}')
    q = '&'.join(filters)
    existing = http_get(f'leads?{q}&select=id,first_name,last_name,company')
    return len(existing) > 0


def normalize_status(s):
    """Map free-form spreadsheet status text to a known enum value."""
    if not s:
        return 'new'
    key = str(s).strip().lower()
    if key in STATUS_MAP:
        return STATUS_MAP[key]
    # Heuristics for partial matches
    for keyphrase, mapped in STATUS_MAP.items():
        if keyphrase in key:
            return mapped
    print(f'  WARNING: unknown status "{s}" — defaulting to "new"')
    return 'new'


# -----------------------------------------------------------------------------
# 1. Import COI — 8 funded buyers from 2025
# -----------------------------------------------------------------------------

def import_coi():
    print('\n=== COI (2025 buyers) ===')
    wb = openpyxl.load_workbook(COI_FILE, data_only=True)
    ws = wb.active

    inserted, skipped = 0, 0
    # Skip header (rows 0-1); footer is row 10 ('Total Equip sold for 2025')
    for row in ws.iter_rows(min_row=3, max_row=10, values_only=True):
        individual = row[0]
        entity = row[1]
        invoice_price = row[2]
        aggregation_fee_pct = row[3]      # e.g. 0.03
        coi_split_amount = row[7]         # Cochran Capital's cut (the "commission")

        if not individual or 'total' in str(individual).lower():
            continue

        first, last = split_name(individual)
        company = entity

        if lead_exists(first, last, company):
            print(f'  SKIP (exists): {individual} / {entity}')
            skipped += 1
            continue

        payload = {
            'first_name': first,
            'last_name': last,
            'company': company,
            'status': 'operating',  # they bought equipment, it's running
            'estimated_equipment_value': float(invoice_price) if invoice_price else None,
            'estimated_total_commission': float(coi_split_amount) if coi_split_amount else None,
            'import_source': 'spreadsheet_2025_coi',
            'notes': f'2025 funded buyer. Equipment: ${invoice_price:,.0f}. '
                     f'Aggregation fee: {aggregation_fee_pct*100:.1f}%. '
                     f'Cochran Capital share: ${coi_split_amount:,.0f}.'
                     if invoice_price and aggregation_fee_pct else None,
        }
        # Use a recent past date so they show up properly on "Joined" sorts
        # (Supabase generates created_at automatically; we override here)
        payload['created_at'] = '2025-12-31T00:00:00Z'

        print(f'  INSERT: {individual} / {entity} -> operating, ${invoice_price:,.0f}')
        result = http_post('leads', [payload])
        if result is not None:
            inserted += 1

    print(f'COI: inserted {inserted}, skipped {skipped}')


# -----------------------------------------------------------------------------
# 2. Import Lead Tracker — active pipeline + historicals
# -----------------------------------------------------------------------------

def import_lead_tracker():
    print('\n=== Lead Tracker (active pipeline) ===')
    wb = openpyxl.load_workbook(LEAD_FILE, data_only=True)
    ws = wb['Cochran Capital Lead Tracker']

    inserted, skipped, skipped_blank = 0, 0, 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: a, Lead Name, Company/LLC, Status, Action Step, Referrer, Notes,
        #          Trust, Intro Email Sent, Referrals From, Invst $ Est,
        #          Yearly Income Est, Slug 1 Cash $, Slug 1 Equip $
        lead_id = row[0]   # e.g. 'L001'
        lead_name = row[1]
        company = row[2]
        status_text = row[3]
        action_step = row[4]
        notes_text = row[6]
        intro_email_sent = row[8]
        invst_est = row[10]
        slug_equip = row[13]

        if not lead_name:
            skipped_blank += 1
            continue

        first, last = split_name(lead_name)

        # Skip COI buyers already imported (avoid dupes)
        if lead_exists(first, last, company):
            print(f'  SKIP (exists): {lead_name}')
            skipped += 1
            continue

        status = normalize_status(status_text)

        # Build combined notes from notes + action_step
        notes_parts = []
        if notes_text and str(notes_text).strip():
            notes_parts.append(str(notes_text).strip())
        if action_step and str(action_step).strip():
            notes_parts.append(f'Next action: {str(action_step).strip()}')
        notes = '\n\n'.join(notes_parts) if notes_parts else None

        # Equipment target — prefer slug_equip, fall back to invst_est
        equip_value = None
        if slug_equip and isinstance(slug_equip, (int, float)) and slug_equip > 0:
            equip_value = float(slug_equip)
        elif invst_est and isinstance(invst_est, (int, float)) and invst_est > 0:
            equip_value = float(invst_est)

        # created_at — prefer intro email date, fall back to now
        created_at = None
        if intro_email_sent and isinstance(intro_email_sent, datetime):
            created_at = intro_email_sent.isoformat() + 'Z'

        payload = {
            'first_name': first,
            'last_name': last,
            'company': str(company).strip() if company else None,
            'status': status,
            'estimated_equipment_value': equip_value,
            'notes': notes,
            'import_source': 'spreadsheet_2026_pipeline',
        }
        if created_at:
            payload['created_at'] = created_at

        print(f'  INSERT: {lead_name} ({status}) {f"-> ${equip_value:,.0f}" if equip_value else ""}')
        result = http_post('leads', [payload])
        if result is not None:
            inserted += 1

    print(f'Lead Tracker: inserted {inserted}, skipped (existing) {skipped}, skipped (blank rows) {skipped_blank}')


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

if __name__ == '__main__':
    print(f'Importing to {SUPABASE_URL}')
    print('(Idempotent — re-running won\'t create duplicates by name match)')

    import_coi()
    import_lead_tracker()

    print('\nDone. Open /admin/leads and click the "Clients" bucket tab to see the 8 buyers,')
    print('then "Active" / "Not now" / "Archived" tabs for the 2026 pipeline.')
